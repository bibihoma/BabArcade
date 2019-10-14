


from __future__ import print_function
import pickle
import os.path
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request

from openpyxl.utils import get_column_letter
import datetime

# If modifying these scopes, delete the file token.pickle.
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']

# The ID and range of a sample spreadsheet.
ELO_SPREADSHEET_ID = '1qkLAN5nPOCeX0d2_I0s23TOV-a2p2DV5R-LQia711FA'
GameResultsRange = 'Fusball!2:6'

#SAMPLE_RANGE_NAME = 'BabyPlayers'
def getGoogleservice():
    """Shows basic usage of the Sheets API.
    Prints values from a sample spreadsheet.
    """
    creds = None
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)

    service = build('sheets', 'v4', credentials=creds)

    return service

def uploadResults(w1,w2,l1,l2,scorerlist = None):
    service = getGoogleservice()
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=ELO_SPREADSHEET_ID,
                                range=GameResultsRange).execute()
    values = result.get('values', [])
    if not values:
        return 'No data found.'

    nbColumns = len(values[0])
    targetRange=get_column_letter(nbColumns+1)+"2:"+get_column_letter(nbColumns+1)+"6"

    now = datetime.datetime.now()

    values = [[now.strftime("%Y-%m-%d")],[w1],[w2],[l1],[l2]]
    body = {'values': values}
    result = service.spreadsheets().values().append(
        spreadsheetId=ELO_SPREADSHEET_ID,
        range=targetRange,
        valueInputOption='USER_ENTERED',
        body = body).execute()

    if scorerlist:
        values = []
        for scorer in scorerlist:
            values.append([scorer])
        targetRange = get_column_letter(nbColumns + 1) + "40:" + get_column_letter(nbColumns + 1) + str(40+len(scorerlist))
        body = {'values': values}
        result = service.spreadsheets().values().append(
            spreadsheetId=ELO_SPREADSHEET_ID,
            range=targetRange,
            valueInputOption='USER_ENTERED',
            body=body).execute()



def main():
    uploadResults('Thomas','Etienne','Jerem','Colin',['Jerem','Thomas','Etienne',"Colin","Jerem"])
    return

if __name__ == '__main__':
    main()
